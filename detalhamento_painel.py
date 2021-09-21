import pandas as pd
from shutil import copy
import os
from openpyxl import load_workbook
import re


source_filename = "U:\\Engenharia\\Usuários\\Guilherme.Kubo\\0.IMPORTAR LISTA DE MATERIAIS\\TESTE 2.xls"

df = pd.read_excel(source_filename, sheet_name='TABELA FINAL')

df['b'] = df['DESCRIÇÃO'].str.findall('\[.*?\]')

df['b'] = df['b'].str.join(', ')
df['b'] = df['b'].str.strip('[]')
print(df)

OV = 'TESTE_ENG'

template_filename = "U:\\Engenharia\\Usuários\\Guilherme.Kubo\\0.IMPORTAR LISTA DE MATERIAIS\\Tabela Estrutura Paineis BR.xlsx"
template_destination = "U:\\Engenharia\\Usuários\\patrick.vieira"

copy(template_filename, template_destination)

template_destination_OV = template_destination+'\\Tabela Estrutura Paineis - ' + OV + '.xlsx'
if os.path.exists(template_destination_OV):
    os.remove(template_destination_OV)

os.rename(template_destination+'\\Tabela Estrutura Paineis BR.xlsx',template_destination_OV)

wb = load_workbook(template_destination_OV)
ws = wb['data']
for i in range(len(df['b'])):
    j = i + 2
    ws[str('G'+str(j))] = df['b'][i]
    ws[str('H'+str(j))] = df['QTD'][i]
    if re.findall('ENG',df['b'][i]):
        ws[str('F'+str(j))] = str(OV)
    i += 1

ws.delete_rows(len(df['b'])+2, 38-len(df['b']))

wb.save(template_destination_OV)


